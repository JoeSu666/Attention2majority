import tensorflow as tf
import numpy as np
import h5py
import os
from mlxtend.plotting import plot_confusion_matrix
from sklearn.metrics import confusion_matrix
import matplotlib.pyplot as plt
from matplotlib.pyplot import MultipleLocator
import openpyxl
import random



def sortkey(e):
    return e[0]


def load_discimage(patch,label, n_class=3):
    if n_class==3:
        if label==1 or label==2:
            label=1
        elif label==3:
            label=2      # if n_class==3, merge class 1 and 2 into 1
    label=tf.one_hot(label,n_class)

    patch=tf.cast(patch,tf.float32)
    patch=tf.keras.applications.resnet.preprocess_input(tf.transpose(patch,(1,2,0)))


    return (patch,label)

def load_image(patch):
    if patch.shape[0]==3:
        patch=np.resize(patch,(1,3,128,128))
    patch=tf.cast(patch,tf.float32)
    patch=tf.keras.applications.resnet.preprocess_input(tf.transpose(patch,(0,2,3,1)))
    #patch=tf.cast(patch,tf.uint8)

    return patch #load image to embed whole dataset

def write_xls(path, history, sheet_name):
    col = 2
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.create_sheet()
    sheet.title = sheet_name
    for i, a in enumerate(history.history["accuracy"]):
        sheet.cell(i+1, 1, a)
    for i, a in enumerate(history.history['val_'+'accuracy']):
        sheet.cell(i+1, 2, a)
    for i, a in enumerate(history.history["loss"]):
        sheet.cell(i+1, 3, a)
    for i, a in enumerate(history.history['val_'+'loss']):
        sheet.cell(i+1, 4, a)
    workbook.save(filename=path)
    print('sheet saved')
    
def kfold_on_class(dic, k, start):
    if start+k > len(dic):
        if start < len(dic):
            test = dic[-k:]
            train = dic[:-k]
        else:
            start = random.choice(range(len(dic)-k+1))
            test = dic[start:start+k]
    else:
        test = dic[start:start+k] 
        train = dic[:start]+dic[start+k:]
    return train, test

def kfold_namelist(dic, n_class, k, start, max_start = 14):
    if start > max_start:
        raise IndexError('Start index should be lower than {}'.format(max_start+1))
    trainnamelist = []
    testnamelist = []
    for l in range(n_class):
        if l+1==3:
            train, test = kfold_on_class(dic[l+1], 9, start)
        else:
            train, test = kfold_on_class(dic[l+1], k, start)
        trainnamelist.extend(train)
        testnamelist.extend(test)
    return trainnamelist, testnamelist

def load_data(bag,label,n_class=3):
    if n_class==3:
        if label==1 or label==2:
            label=1
        elif label==3:
            label=2      # if n_class==3, merge class 1 and 2 into 1

    label=tf.one_hot(label,n_class)
    
    return (bag,label)


def plot_metric(history, metric, K=None, start=None, save_dir=None):
    train_metrics = history.history[metric]
    val_metrics = history.history['val_'+metric]
    epochs = range(1, len(train_metrics) + 1)
    if metric=='accuracy':  
        x_major_locator=MultipleLocator(20)
        y_major_locator=MultipleLocator(0.1)
        ax=plt.gca()
        ax.xaxis.set_major_locator(x_major_locator)
        ax.yaxis.set_major_locator(y_major_locator)
        plt.axis([0,150,0,1.1])
    plt.grid(linestyle='-.')
    plt.plot(epochs, train_metrics, 'b.--')
    plt.plot(epochs, val_metrics, 'r.-')
    plt.title('Training and validation '+ metric)
    plt.xlabel("Epochs")
    plt.ylabel(metric)
    plt.legend(["train_"+metric, 'val_'+metric])
    if save_dir:        
        plt.savefig(save_dir+str(K)+'sample'+str(start)+'fold'+metric+'.jpg')
    plt.show()
    
def chunks(l,n):
    for i in range(0,len(l),n):
        yield l[i:i+n]
        