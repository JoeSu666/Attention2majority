import tensorflow as tf
import numpy as np
import h5py
import os
from mlxtend.plotting import plot_confusion_matrix
from sklearn.metrics import confusion_matrix
import matplotlib.pyplot as plt
from matplotlib.pyplot import MultipleLocator
import openpyxl
import random
from tensorflow.keras import backend



def sortkey(e):
    return e[0]


def load_discimage(patch,label):
    if label == 0:
        label = tf.constant(0, tf.int8)
    else:
        label = tf.constant(1, tf.int8)
        
    patch=tf.cast(patch,tf.float32)
    patch=tf.keras.applications.resnet.preprocess_input(tf.transpose(patch,(1,2,0)))


    return (patch,label)

def load_image(patch):
    if patch.shape[0]==3:
        patch=np.resize(patch,(1,3,128,128))
    patch=tf.cast(patch,tf.float32)
    patch=tf.keras.applications.resnet.preprocess_input(tf.transpose(patch,(0,2,3,1)))
    #patch=tf.cast(patch,tf.uint8)

    return patch #load image to embed whole dataset

def write_xls(path, history, sheet_name):
    col = 2
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.create_sheet()
    sheet.title = sheet_name
    for i, a in enumerate(history.history["accuracy"]):
        sheet.cell(i+1, 1, a)
    for i, a in enumerate(history.history['val_'+'accuracy']):
        sheet.cell(i+1, 2, a)
    for i, a in enumerate(history.history["loss"]):
        sheet.cell(i+1, 3, a)
    for i, a in enumerate(history.history['val_'+'loss']):
        sheet.cell(i+1, 4, a)
    workbook.save(filename=path)
    print('sheet saved')
    
def kfold_on_class(dic, k, start):
    if start+k > len(dic):
        if start < len(dic):
            test = dic[-k:]
            train = dic[:-k]
        else:
            start = random.choice(range(len(dic)-k+1))
            test = dic[start:start+k]
    else:
        test = dic[start:start+k] 
        train = dic[:start]+dic[start+k:]
    return train, test

def kfold_namelist(dic, n_class, k, start, max_start = 14):
    if start > max_start:
        raise IndexError('Start index should be lower than {}'.format(max_start+1))
    trainnamelist = []
    testnamelist = []
    for l in range(n_class):
        if l+1==3:
            train, test = kfold_on_class(dic[l+1], 9, start)
        else:
            train, test = kfold_on_class(dic[l+1], k, start)
        trainnamelist.extend(train)
        testnamelist.extend(test)
    return trainnamelist, testnamelist

def load_data(bag,label):
    if label == 0:
        label = tf.constant(0, tf.int32)
    else:
        label = tf.constant(1, tf.int32)
    bag = tf.cast(bag, tf.float32)
    
    return (bag,label)


def plot_metric(history, metric, K=None, start=None, save_dir=None):
    train_metrics = history.history[metric]
    val_metrics = history.history['val_'+metric]
    epochs = range(1, len(train_metrics) + 1)
    if metric=='accuracy':  
        x_major_locator=MultipleLocator(20)
        y_major_locator=MultipleLocator(0.1)
        ax=plt.gca()
        ax.xaxis.set_major_locator(x_major_locator)
        ax.yaxis.set_major_locator(y_major_locator)
        plt.axis([0,200,0,1.1])
    plt.grid(linestyle='-.')
    plt.plot(epochs, train_metrics, 'b.--')
    plt.plot(epochs, val_metrics, 'r.-')
    plt.title('Training and validation '+ metric)
    plt.xlabel("Epochs")
    plt.ylabel(metric)
    plt.legend(["train_"+metric, 'val_'+metric])
    if save_dir:        
        plt.savefig(save_dir+str(K)+'sample'+str(start)+'fold'+metric+'.jpg')
    plt.show()
    
def chunks(l,n):
    for i in range(0,len(l),n):
        yield l[i:i+n]
        
def sensitivity(y_true, y_pred):
    true_positives = backend.sum(backend.round(backend.clip(y_true * y_pred, 0, 1)))
    possible_positives = backend.sum(backend.round(backend.clip(y_true, 0, 1)))
    return true_positives / (possible_positives + backend.epsilon())

def specificity(y_true, y_pred):
    true_negatives = backend.sum(backend.round(backend.clip((1-y_true) * (1-y_pred), 0, 1)))
    possible_negatives = backend.sum(backend.round(backend.clip(1-y_true, 0, 1)))
    return true_negatives / (possible_negatives + backend.epsilon())
    
def repeatpad(length, data, sortidxs, encoded_shape):
    '''
    repeatedly pad the data
    '''
    datalength = data.shape[0]
    embedding = np.empty((0, encoded_shape))
    print(embedding.shape)
    padtime = length//datalength
    remainder = length%datalength
    
    # embedding = np.concatenate((embedding, np.zeros((length, encoded_shape))), axis=0)
    for i in range(padtime):
        embedding = np.concatenate((embedding, data[sortidxs]), axis=0)
    embedding = np.concatenate((embedding, data[sortidxs[:remainder]]), axis=0)
    
    return embedding
    
    
    
    
    